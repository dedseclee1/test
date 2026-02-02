# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
import pyodbc
import traceback
from pathlib import Path
import datetime
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment

# ============== 用户配置区 ==============
DB_CONN_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=192.168.0.117;"
    "DATABASE=FQD;"
    "UID=zhitan;"
    "PWD=Zt@forcome;"
)

# 截图中的关键配置
ROW_IDX_HEADER_MAIN = 2    # 主表头所在行 (如：工单单号、车间)
ROW_IDX_HEADER_DATE = 3    # 日期表头所在行 (如：1/25, 1/26)
ROW_IDX_DATA_START = 4     # 数据起始行

COL_NAME_WORKSHOP = "车间"      # 截图K列
COL_NAME_WO_TYPE = "单别"       # 截图E列
COL_NAME_WO_NO = "工单单号"     # 截图D列
COL_NAME_WO_TOTAL = "工单预计生产总量" # 需确认ERP字段或Excel是否有此列，如果没有，代码将从ERP查询

# ============== 应用程序类 ==============

class DailyPlanAvailabilityApp:
    def __init__(self, root):
        self.root = root
        self.root.title("每日排程齐套分析工具 v4.0 (横向日期版)")
        self.root.geometry("900x650")

        # 样式定义
        self.red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # 变量绑定
        self.file_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.selected_date_str = tk.StringVar()
        self.selected_workshop = tk.StringVar()

        # 缓存数据
        self.date_column_map = {}  # {'2026-01-25': 105, ...} 存储日期对应的列号
        self.all_workshops = []
        self.col_map_main = {}     # 主表头列索引

        self._create_widgets()

    def _create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 文件选择
        file_frame = ttk.LabelFrame(main_frame, text="1. 数据源", padding="5")
        file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(file_frame, textvariable=self.file_path, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="浏览Excel...", command=self._select_file).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(file_frame, text="   工作表:").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(file_frame, textvariable=self.sheet_name, state="disabled", width=15)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_selected)

        # 2. 筛选设置
        filter_frame = ttk.LabelFrame(main_frame, text="2. 计划筛选 (自动扫描第3行日期)", padding="10")
        filter_frame.pack(fill=tk.X, pady=5)

        ttk.Label(filter_frame, text="选择日期:").grid(row=0, column=0, sticky="w")
        self.date_combo = ttk.Combobox(filter_frame, textvariable=self.selected_date_str, state="disabled", width=25)
        self.date_combo.grid(row=0, column=1, padx=5, sticky="w")

        ttk.Label(filter_frame, text="选择车间:").grid(row=0, column=2, sticky="w", padx=(20, 0))
        self.workshop_combo = ttk.Combobox(filter_frame, textvariable=self.selected_workshop, state="disabled", width=20)
        self.workshop_combo.grid(row=0, column=3, padx=5, sticky="w")

        # 3. 操作区
        action_frame = ttk.LabelFrame(main_frame, text="3. 执行", padding="10")
        action_frame.pack(fill=tk.X, pady=10)
        
        btn = ttk.Button(action_frame, text="计算齐套率 -> 写入A列 -> 标红缺料", command=self._run_analysis)
        btn.pack(fill=tk.X, padx=100)

        # 4. 日志
        self.log_text = tk.Text(main_frame, height=12, state="disabled", font=("Consolas", 9), bg="#F0F0F0")
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=5)

    def _log(self, msg):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        self.root.update_idletasks()

    def _select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.file_path.set(path)
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                self.sheet_combo['values'] = wb.sheetnames
                if wb.sheetnames: 
                    self.sheet_combo.current(0)
                    self._on_sheet_selected(None)
                self.sheet_combo.config(state="readonly")
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件: {e}")

    def _on_sheet_selected(self, event):
        """当Sheet改变时，扫描表头结构"""
        file_path = self.file_path.get()
        sheet_name = self.sheet_name.get()
        if not file_path or not sheet_name: return

        self._log("正在扫描Excel结构 (表头和日期列)...")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]

            # 1. 扫描主表头 (第2行)
            self.col_map_main = {}
            for cell in ws[ROW_IDX_HEADER_MAIN]:
                if cell.value:
                    self.col_map_main[str(cell.value).strip()] = cell.column

            # 检查关键列是否存在
            required_cols = [COL_NAME_WORKSHOP, COL_NAME_WO_TYPE, COL_NAME_WO_NO]
            missing = [c for c in required_cols if c not in self.col_map_main]
            if missing:
                messagebox.showwarning("警告", f"未找到关键列: {missing}\n请检查Excel表头(第2行)名称是否与配置一致。")
                return

            # 2. 扫描日期列 (第3行)
            self.date_column_map = {}
            # 假设日期列在主数据列之后，或者遍历整行。
            # 通常排程表的日期在后面，我们遍历所有有值的单元格
            for cell in ws[ROW_IDX_HEADER_DATE]:
                val = cell.value
                date_str = self._parse_excel_date(val)
                if date_str:
                    self.date_column_map[date_str] = cell.column

            dates = sorted(list(self.date_column_map.keys()))
            if not dates:
                self._log("警告: 在第3行未找到任何日期格式的表头！")
            else:
                self._log(f"找到 {len(dates)} 个排程日期: {dates[0]} 至 {dates[-1]}")
                self.date_combo['values'] = dates
                self.date_combo.current(0)
                self.date_combo.config(state="readonly")

            # 3. 扫描车间 (读取数据行去重)
            col_ws_idx = self.col_map_main[COL_NAME_WORKSHOP]
            workshops = set()
            # 简单扫描前1000行或全部
            for row in ws.iter_rows(min_row=ROW_IDX_DATA_START, min_col=col_ws_idx, max_col=col_ws_idx, values_only=True):
                if row[0]: workshops.add(str(row[0]).strip())
            
            ws_list = sorted(list(workshops))
            self.workshop_combo['values'] = ["全部车间"] + ws_list
            self.workshop_combo.current(0)
            self.workshop_combo.config(state="readonly")

        except Exception as e:
            traceback.print_exc()
            self._log(f"扫描失败: {e}")

    def _parse_excel_date(self, val):
        """解析单元格中的日期 (支持 Excel日期数字, datetime对象, 字符串)"""
        if val is None: return None
        try:
            dt = None
            if isinstance(val, (datetime.datetime, datetime.date)):
                dt = val
            elif isinstance(val, (int, float)):
                # Excel serial date
                dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(val))
            elif isinstance(val, str):
                # 尝试解析字符串 "1/25", "2026/1/25"
                # 根据截图，可能是不带年份的 "M/D"
                parts = val.strip().split('/')
                if len(parts) == 2: # M/D
                    # 这是一个模糊日期，默认用当前年份或下一年？
                    # 截图里有2025和2026，为了保险，最好Excel里是真日期格式。
                    # 这里暂且假设是真日期格式被读取。如果是纯文本，需要额外逻辑。
                    # 临时逻辑：如果解析出纯文本，返回原文本
                    return val.strip() 
                elif len(parts) == 3:
                     dt = datetime.datetime.strptime(val.strip(), "%Y/%m/%d")
            
            if dt:
                return dt.strftime("%Y-%m-%d")
            return None
        except:
            return None

    def _run_analysis(self):
        target_date = self.selected_date_str.get()
        target_workshop = self.selected_workshop.get()
        file_path = self.file_path.get()
        sheet_name = self.sheet_name.get()

        if not target_date or not self.date_column_map:
            messagebox.showwarning("提示", "请先选择一个有效的日期。")
            return

        target_col_idx = self.date_column_map.get(target_date)
        if not target_col_idx:
            self._log(f"错误: 无法找到日期 {target_date} 对应的列号。")
            return

        if not messagebox.askyesno("确认", 
            f"目标日期: 【{target_date}】\n"
            f"目标车间: 【{target_workshop}】\n\n"
            f"程序将读取Excel中对应日期的列(排产数)，\n"
            "并重新计算齐套率写入【A列】。\n"
            "这会覆盖A列原有内容，是否继续？"):
            return

        try:
            self._log("="*50)
            self._log(f"开始分析... 目标列号: {target_col_idx}")
            
            # 1. 读取符合条件的计划行
            plans = self._load_daily_plans(file_path, sheet_name, target_col_idx, target_workshop)
            
            if not plans:
                self._log("该日期/车间下没有非零的排产计划。")
                messagebox.showinfo("无数据", "该日期下没有找到排产数量 > 0 的计划。")
                return
            
            self._log(f"找到 {len(plans)} 条有效计划。")

            # 2. 获取数据库数据
            wo_keys = list(set(p['wo_key'] for p in plans))
            self._log("查询ERP工单BOM...")
            wo_details = self._fetch_erp_data(wo_keys)
            
            all_parts = set()
            for w in wo_details.values():
                for b in w['bom']: all_parts.add(b['part'])
            
            self._log(f"查询ERP实时库存 ({len(all_parts)} 种物料)...")
            inventory = self._fetch_inventory(list(all_parts))

            # 3. 模拟计算
            self._log("执行齐套模拟 (All-or-Nothing)...")
            results, stats = self._simulate(plans, wo_details, inventory)

            # 4. 写入Excel
            self._log("写入结果到Excel...")
            self._write_excel(file_path, sheet_name, results)

            # 5. 结果弹窗
            self._show_summary(target_date, stats)

        except Exception as e:
            traceback.print_exc()
            self._log(f"严重错误: {e}")
            messagebox.showerror("错误", str(e))

    def _load_daily_plans(self, file_path, sheet_name, date_col_idx, filter_ws):
        """读取指定列有数值的行"""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        # 获取列索引
        c_ws = self.col_map_main[COL_NAME_WORKSHOP]
        c_type = self.col_map_main[COL_NAME_WO_TYPE]
        c_no = self.col_map_main[COL_NAME_WO_NO]
        
        plans = []
        
        # 遍历数据行
        # 优化：只读取必要的列。但 openpyxl read_only 只能按行读取。
        for row in ws.iter_rows(min_row=ROW_IDX_DATA_START):
            # 获取该行日期列的值 (注意 iter_rows 返回的是元组，索引从0开始)
            # date_col_idx 是 1-based，所以要 -1
            try:
                # 检查索引是否越界
                if date_col_idx > len(row): continue
                
                daily_qty_val = row[date_col_idx - 1].value
                
                # 检查是否有排产 (非空 且 > 0)
                # 处理 Excel 可能的括号负数或文本格式，截图里 "(30)" 可能是负数或者是备注
                # 截图显示正数如 "65", "9", "30"。
                if isinstance(daily_qty_val, (int, float)) and daily_qty_val > 0:
                    
                    # 检查车间
                    curr_ws = row[c_ws - 1].value
                    curr_ws = str(curr_ws).strip() if curr_ws else "未分类"
                    
                    if filter_ws != "全部车间" and curr_ws != filter_ws:
                        continue
                        
                    # 提取工单信息
                    wo_type = row[c_type - 1].value
                    wo_no = row[c_no - 1].value
                    
                    if wo_type and wo_no:
                        plans.append({
                            'row_idx': row[0].row, # 获取真实行号
                            'wo_key': (str(wo_type).strip(), str(wo_no).strip()),
                            'daily_qty': float(daily_qty_val),
                            'workshop': curr_ws
                        })
            except IndexError:
                continue
                
        return plans

    def _fetch_erp_data(self, wo_keys):
        """查询ERP BOM和工单总量"""
        if not wo_keys: return {}
        
        # 构造批量查询
        # 注意：如果工单太多，可能需要分批。这里简化处理。
        conditions = []
        for t, n in wo_keys:
            conditions.append(f"(TA.TA001='{t}' AND TA.TA002='{n}')")
        
        if not conditions: return {}
        
        where_sql = " OR ".join(conditions)
        
        sql = f"""
            SELECT 
                RTRIM(TA.TA001) as ta001, RTRIM(TA.TA002) as ta002, 
                TA.TA015 as wo_total_qty,
                RTRIM(TB.TB003) as part_no, ISNULL(RTRIM(MB.MB002),'') as part_name,
                TB.TB004 as req_qty, TB.TB005 as iss_qty
            FROM MOCTA TA
            INNER JOIN MOCTB TB ON TA.TA001 = TB.TB001 AND TA.TA002 = TB.TB002
            LEFT JOIN INVMB MB ON TB.TB003 = MB.MB001
            WHERE {where_sql}
        """
        
        try:
            with pyodbc.connect(DB_CONN_STRING) as conn:
                df = pd.read_sql(sql, conn)
        except Exception as e:
            raise Exception(f"数据库查询失败: {e}")

        data = defaultdict(lambda: {'total': 0, 'bom': []})
        for _, row in df.iterrows():
            k = (row['ta001'], row['ta002'])
            data[k]['total'] = float(row['wo_total_qty'])
            data[k]['bom'].append({
                'part': row['part_no'],
                'name': row['part_name'],
                'req': float(row['req_qty']),
                'iss': float(row['iss_qty'])
            })
        return data

    def _fetch_inventory(self, parts):
        if not parts: return {}
        # 简单处理，如果parts太多可能报错，实际生产建议分批
        p_str = ",".join(f"'{p}'" for p in parts)
        sql = f"SELECT RTRIM(MC001) as p, SUM(MC007) as q FROM INVMC WHERE MC001 IN ({p_str}) GROUP BY MC001"
        try:
            with pyodbc.connect(DB_CONN_STRING) as conn:
                df = pd.read_sql(sql, conn)
            return pd.Series(df.q.values, index=df.p).to_dict()
        except:
            return {}

    def _simulate(self, plans, wo_data, inventory):
        """
        plans: list of dict {'row_idx', 'wo_key', 'daily_qty', 'workshop'}
        """
        running_inv = inventory.copy()
        results = {} # {row_idx: {'text': str, 'short': bool}}
        
        stats = {
            'total_plans': 0,
            'kitted_plans': 0,
            'ws_stats': defaultdict(lambda: {'total': 0, 'kitted': 0})
        }
        
        for p in plans:
            row = p['row_idx']
            wo_key = p['wo_key']
            daily_qty = p['daily_qty'] # 本日计划数
            ws = p['workshop']
            
            stats['total_plans'] += 1
            stats['ws_stats'][ws]['total'] += 1
            
            info = wo_data.get(wo_key)
            if not info or not info['bom']:
                results[row] = {'text': "无ERP信息", 'short': True}
                continue
                
            wo_total_qty = info['total'] # 工单总排产量
            
            # --- 核心逻辑 ---
            # 净需求 = Min(工单剩余未发, 本次计划对应的理论配比需求)
            
            items_needed = 0
            items_kitted = 0
            shortage_details = [] # ["品名(缺xx)"]
            
            # 记录本单如果要扣减库存，需要扣多少
            to_deduct = {} # {part: qty}
            
            is_fully_kitted = True
            
            # 最小可产套数计算 (受限于短板物料)
            min_possible_sets = 9999999
            
            for bom in info['bom']:
                part = bom['part']
                remain_issue = max(0, bom['req'] - bom['iss']) # 剩余未领
                
                # 理论配比需求 = (单耗) * 本次计划数
                unit_usage = bom['req'] / wo_total_qty if wo_total_qty > 0 else 0
                theo_demand = daily_qty * unit_usage
                
                # 实际净需求
                net_demand = min(remain_issue, theo_demand)
                
                if net_demand <= 0: continue # 不需要领料
                
                items_needed += 1
                to_deduct[part] = net_demand
                
                current_stock = running_inv.get(part, 0)
                
                # 计算该物料支持生产多少个成品
                can_make = int(current_stock // unit_usage) if unit_usage > 0 else 999999
                min_possible_sets = min(min_possible_sets, can_make)
                
                if current_stock >= net_demand:
                    items_kitted += 1
                else:
                    is_fully_kitted = False
                    short_qty = net_demand - current_stock
                    shortage_details.append(f"{bom['name']}(缺{short_qty:.1f})")

            # 修正最小可产套数 (不能超过计划数)
            actual_possible_sets = min(int(daily_qty), min_possible_sets)
            if items_needed == 0: 
                actual_possible_sets = int(daily_qty)
                is_fully_kitted = True

            # 计算齐套率
            rate = (items_kitted / items_needed) if items_needed > 0 else 1.0
            
            # 构造A列文本
            txt = f"齐套率:{rate:.0%} | 可产:{actual_possible_sets}"
            if not is_fully_kitted:
                txt += " | 缺:" + ",".join(shortage_details[:2]) # 只显示前2个缺料避免太长
                if len(shortage_details) > 2: txt += "..."
            
            results[row] = {'text': txt, 'short': not is_fully_kitted}
            
            # 库存扣减 (All-or-Nothing)
            if is_fully_kitted:
                stats['kitted_plans'] += 1
                stats['ws_stats'][ws]['kitted'] += 1
                for part, qty in to_deduct.items():
                    running_inv[part] -= qty
            
        return results, stats

    def _write_excel(self, file_path, sheet_name, results):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        font = Font(name="微软雅黑", size=9)
        align = Alignment(wrap_text=True, vertical="center")
        
        # 为了不影响未筛选的行，我们只更新 results 中存在的行
        # 如果需要清空A列其他行，需要遍历整个A列。
        # 建议：仅覆盖计算的行，这样保留历史记录？或者根据需求清空。
        # 这里逻辑：只更新计算行。
        
        for row_idx, res in results.items():
            cell = ws.cell(row=row_idx, column=1) # A列
            cell.value = res['text']
            cell.font = font
            cell.alignment = align
            
            if res['short']:
                cell.fill = self.red_fill
            else:
                cell.fill = self.green_fill # 全齐套标绿，方便区分

        wb.save(file_path)

    def _show_summary(self, date_str, stats):
        msg = f"日期: {date_str}\n"
        msg += f"总计划行数: {stats['total_plans']}\n"
        msg += f"全齐套行数: {stats['kitted_plans']}\n"
        total_rate = stats['kitted_plans'] / stats['total_plans'] if stats['total_plans'] else 0
        msg += f"总体齐套率: {total_rate:.1%}\n\n"
        msg += "--- 各车间详情 ---\n"
        
        for ws, d in stats['ws_stats'].items():
            r = d['kitted'] / d['total'] if d['total'] else 0
            msg += f"{ws}: {r:.1%} ({d['kitted']}/{d['total']})\n"
            
        messagebox.showinfo("分析结果", msg)

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = DailyPlanAvailabilityApp(root)
        root.mainloop()
    except Exception as e:
        # 如果报错，提供一个简单的弹窗
        import tkinter.messagebox
        tkinter.messagebox.showerror("启动失败", str(e))